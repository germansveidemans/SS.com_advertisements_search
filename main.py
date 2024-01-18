import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook, load_workbook
import pandas

# Ievieto datus ss.com šūnās no options.csv faila, lai atrastu sludinājumus.
def Fill_cells(options, driver, i):
    
    #  Saprotama datu apstrāde pa mainīgājiem
    option_cena = options[i][0].split("-")
    option_gads = options[i][1].split("-")
    option_tilpums = options[i][2]
    
    if option_tilpums.find("-") != 0:
        option_tilpums = option_tilpums.split("-")
        
    option_dzinejs = options[i][3]
    option_atrkarba = options[i][4]
    option_virstips = options[i][5]
    option_krasa = options[i][6]
    
    # Datu ievietošana šūnās
    find = driver.find_element(By.ID, "mtd_97")
    find.click()
    
    find = driver.find_element(By.ID, "f_o_8_min")
    find.send_keys(option_cena[0])
    find = driver.find_element(By.ID, "f_o_8_max")
    find.send_keys(option_cena[1])
    
    find = driver.find_element(By.ID, "f_o_18_min")
    find.send_keys(option_gads[0])
    find = driver.find_element(By.ID, "f_o_18_max")
    find.send_keys(option_gads[1])

    find = driver.find_element(By.ID, "f_o_15_min")
    find.send_keys(option_tilpums[0])
    find = driver.find_element(By.ID, "f_o_15_max")
    if len(option_tilpums) > 1:
        find.send_keys(option_tilpums[1])
    else:
        find.send_keys(option_tilpums[0])
    
    find = driver.find_element(By.ID, "f_o_34")
    find.send_keys(option_dzinejs)
    
    find = driver.find_element(By.ID, "f_o_35")
    find.send_keys(option_atrkarba)
    
    find = driver.find_element(By.ID, "f_o_32")
    find.send_keys(option_virstips)
    
    if option_krasa != "-":
        find = driver.find_element(By.ID, "f_o_17")
        find.send_keys(option_krasa)
    
    find = driver.find_element(By.LINK_TEXT, "Nobrauk.")
    find.click()
    return

# Apstradā sludinājumu informāciju un pierāksta to tabulā.
def Find_information(driver, number, result_wb):
    
    sludinajumi_id = []
    saites = []
    num = -1
    # Informāciju meklēšana par marku un modeli, gadu, cenu, dzinēju tilpumu un saiti.
    sl_class = driver.find_elements(By.CLASS_NAME, "msga2-o")
    
    for i in sl_class:
        info = i.text
        if info != "-":
            sludinajumi_id.append(info)
        hh = i.find_element(By.TAG_NAME, 'a').get_attribute("href")
        num = len(saites)-1
        if num >= 0:
            if hh != saites[num]:
                saites.append(hh)
        else:
            saites.append(hh)
    # Nobraukuma informācijas meklēšana
    nobraukums = []
    
    sl_class = driver.find_elements(By.CLASS_NAME, "msga2-r")
    for i in sl_class:
        info = i.text
        nobraukums.append(info)
        
    # No sarakstā sludinajumi_id kārto pa atsēvišķiem sarakstiem
    marka = []
    
    for i in range(0,len(sludinajumi_id),4):
        marka.append(sludinajumi_id[i])

    gads = []
    
    for i in range(1,len(sludinajumi_id),4):
        gads.append(sludinajumi_id[i])
 
    tilpums = []
    
    for i in range(2,len(sludinajumi_id),4):
        tilpums.append(sludinajumi_id[i])

    cena = []
    
    for i in range(3,len(sludinajumi_id),4):
        cena.append(sludinajumi_id[i])

    # Tabulas izveidošana
    w_results = result_wb.create_sheet("Lapa_" + str(number))
    w_results.append(["Marka","Gads","Dz.Tilpums","Nobraukums", "Cena", "Saite"])
    
    # Gadījumā ja ss.com vietnē nobraukums ir "-", pievieno to pie saraksta
    if len(marka) != len(nobraukums):
        for i in range(len(marka)-len(nobraukums)):
            nobraukums.append("—")
            
    # Ierakstā informāciju tabulā
    for marka, gads, tilpums, nobraukums, cena, saite in zip(marka, gads, tilpums, nobraukums, cena, saites):
        w_results.append([marka, gads, tilpums, nobraukums, cena, saite])
    result_wb.save("result.xlsx")
    
    return

# Datus pievienošana sarakstā
options_csv = pandas.read_csv("options.csv")
options = options_csv.values.tolist()

# ss.com vietnēs atvēršana
service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service = service, options = option)
url = "https://www.ss.com/"
driver.get(url)

result_wb = Workbook()

# Pēc parametru skaita atkarto visas iepriekš minētas darbības
for i in range(len(options)):
    Fill_cells(options, driver, i)
    Find_information(driver, i, result_wb)
    time.sleep(2)
    find = driver.find_element(By.CLASS_NAME, "page_header_logo")
    find.click()
    
driver.quit()